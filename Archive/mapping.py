from openpyxl import load_workbook

path = r'C:\Users\sunsh\Documents\Mapping doc\fesruleId -shawn.xlsx'
wb = load_workbook(filename = path)
ws1 = wb['Sheet']
ws2 = wb['Mapped']

query1 = []
query2 = []
query3 = []
query4 = []
query5 = []
query6 = []

for row in ws1.iter_rows(min_row=1, min_col=1, max_row=ws1.max_row, max_col=1):
    value = row[0].value
    q1 = fr"max(case when a.fesruleid like '{value}' then prvalue end) {value}_ACT,"
    q2 = fr"max(case when a.fesruleid like '{value}' then paramvalue end) {value}_Thresh,"
    q3 = fr"max(case when a.fesruleid like '{value}' then triggered end) {value}_Flag,"
    q4 = fr"b.{value}_ACT"
    q5 = fr"b.{value}_Thresh"
    q6 = fr"b.{value}_Flag"

    query1.append(q1)
    query2.append(q2)
    query3.append(q3)
    query4.append(q4)
    query5.append(q5)
    query6.append(q6)


query = list(zip(query1, query2, query3, query4, query5, query6))
for row in query:
    ws2.append(row)

wb.save(r'C:\Users\sunsh\Documents\Mapping doc\fesruleId -shawn_filled.xlsx')



