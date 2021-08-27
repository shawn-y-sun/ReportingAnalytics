from .metrics import Metrics
from .template import Template, Entry
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import re
import pandas as pd
from tqdm import tqdm
import functools
import os

class Write(Metrics):
    """An obeject that writes the metrics of dataset into a reporting template"""

    def __init__(self, template, dataset, strategy=None, filter=None):
        """Create an Write object

        :param template: a Template object
        :param dataset: a Pandas dataframe
        :param strategy: a Strategy object
        :param filter: a Filter object
        """
        super().__init__(dataset, strategy, filter)

        # Inherit variables from Template
        self.template = template
        self.entries = self.template.entries.entries
        self.file = self.template.file
        self.name = self.template.name
        self.date_format = self.template.date_format
        self.slicers = self.template.slicers
        self.dir = None
    
    @functools.cached_property
    def wb(self):
        return load_workbook(filename=self.file)
    
    @property
    def sheets(self):
        return self.wb.sheetnames

    @property
    def destination(self):
        dir_path = os.getcwd() if self.dir is None else self.dir
        report_name = fr"{dir_path}\{self.name}.xlsx"
        return report_name
    
    def save(self, name=None):
        if name is None:
            self.wb.save(filename = self.destination)
            print(f"!Report saved at {self.destination}")
        else:
            dir_path = os.getcwd() if self.dir is None else self.dir
            temp_filename = fr"{dir_path}\{name}.xlsx"
            self.wb.save(filename = temp_filename)
            print(f"!Report saved at {temp_filename}")

    def write_col(self, entry, month_ind=None, year_ind=None):
        sheet = entry.sheet if isinstance(entry.sheet, str) else \
            self.sheets[entry.sheet - 1]
        cell = entry.cell
        ws = self.wb[sheet]
        name = entry.name

        if year_ind is not None and month_ind is not None:
            val = self.get_yearly_monthly(name, entry.fiscal, entry.mtype)[year_ind][month_ind]
        elif year_ind is not None:
            val = self.get_yearly(name, entry.fiscal, entry.mtype)[year_ind]
        elif month_ind is not None:
            val = self.get_monthly(name, entry.mtype)[month_ind]
        
        elif entry.mtype:
            val = getattr(self, entry.mtype)
            val = val[name] if val is not None else 0
        elif name != 'date':
            val = getattr(self, name)
        

        if entry.mtier is not None:

            mtier = f"Tier {entry.mtier}" if entry.mtier != 'Total' else entry.mtier
            try:
                val = val[mtier].iloc[:,0].values.tolist()
            except:
                val = 0

            
        if name == 'date':
            ws[cell] = f"{self.end_date:{self.date_format}}"
        elif isinstance(val, (int, float)):
            ws[cell] = val
        elif isinstance(val, dict):
            xy = coordinate_from_string(cell)
            row = xy[1]
            for count in val.values():
                if isinstance(count, dict):
                    for value in count.values():
                        try:
                            value = value / sum(count.values()) if entry.distribution else value
                        except ZeroDivisionError:
                            value = 0
                        cell_new = xy[0] + str(row)
                        ws[cell_new] = value
                        row += 1
                else:
                    try:
                        count = count / sum(val.values()) if entry.distribution else count
                    except ZeroDivisionError:
                        count = 0
                    cell_new = xy[0] + str(row)
                    ws[cell_new] = count
                    row += 1
                
        elif isinstance(val, list):
            for index, count in enumerate(val):
                xy = coordinate_from_string(cell)
                row = xy[1] + index
                cell_new = xy[0] + str(row)
                try:
                    count = count / sum(val) if entry.distribution else count
                except ZeroDivisionError:
                    count = 0
                ws[cell_new] = count


    def write_df(self, entry):
        xy = coordinate_from_string(entry.cell)
        col = xy[0]
        col_ind = column_index_from_string(col)
        row = str(xy[1])

        rp_years = self.f_years if entry.fiscal else self.years
        rp_months = self.f_months if entry.fiscal else self.months

        for yr_ind, rp_year in enumerate(rp_years):
    
            yr_pos = 1 if entry.yearly else 0
            mth_len = len(rp_months)
            year_ind = rp_year
            if entry.monthly:
                col_yly_new = (col_ind - 1) + (yr_ind + 1) * (mth_len + yr_pos)
                for i in range(mth_len):
                    col_new = col_ind + (yr_ind * (mth_len + yr_pos)) + i
                    col_new = get_column_letter(col_new)
                    cell_new = col_new + row

                    month_ind = rp_months[i]
                    entry_new = entry.copy()
                    entry_new.cell = cell_new
                    self.write_col(entry_new, month_ind=month_ind, year_ind=year_ind)
            
            else:
                col_yly_new = col_ind + yr_ind
            
            if entry.yearly:
                col_yly_new = get_column_letter(col_yly_new)
                cell_yly_new = col_yly_new + row
                month_ind = None
                entry_yly_new = entry.copy()
                entry_yly_new.cell = cell_yly_new
                self.write_col(entry_yly_new, month_ind=month_ind, year_ind=year_ind)


    
    def write(self, name=None):
        """Write the mertics into the template and save the file

        :param name: the alternative name you want to save the report
        """
        print('Writing File ...')
        for entry in tqdm(self.entries, desc='Writing Entries',
                            unit='entry', leave=False):
            if entry.yearly or entry.monthly:
                self.write_df(entry)
            else:
                self.write_col(entry)

        self.save(name=name)

        if self.slicers is not None:
            print('Adding Slicers ...')
            self.slicers.add(self.destination)
            print('Slicers Added!')
    

    
